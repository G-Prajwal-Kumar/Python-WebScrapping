import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import threading
import time

from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as uReq

from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium import webdriver

import os
import ctypes
ctypes.windll.shcore.SetProcessDpiAwareness(1)








#Tkiner ______Start_______
import tkinter
from tkinter import *
from tkinter.filedialog import askopenfile

class Scrape():
    def __init__(self):
        self.rollNos = []
        self.outs = {}
        self.NameLengths = []
        self.Subss = []
        self.Subs = []
    
    def startReading(self):
        wb1 = openpyxl.load_workbook(Input)
        sheets = wb1.sheetnames
        self.rollNos = []
        for i in range(len(sheets)):
            wb1.active = i
            sheet1 = wb1.active
            row = sheet1.max_row
            roll=[]
            for i in range(2,row+1):
                roll.append(sheet1.cell(i,1).value)
            self.rollNos.append(roll)
        self.outs = {}
        self.NameLengths = []



    def multiThread_RV_(self, i, start, end,):
        NameLength = 0
        s=Service(DriverLocation)
        driver = webdriver.Chrome(service=s)
        driver.get(UrlRv)
        chk = {}
        for q in range(start,end):        
            element = driver.find_elements(By.TAG_NAME, "input")
            if len(element) == 3:
                element[1].send_keys(q)
                element[2].click()
            else:
                element[2].send_keys(q)
                element[3].click()
            
            html = driver.page_source
            pageSoup = soup(html,"html.parser")
            container = pageSoup.find("table",{"id":"AutoNumber4"})
            
            if container != None:
                containerName = pageSoup.find("table",{"id":"AutoNumber3"})
                nameFull = containerName.findAll("tr")
                name = nameFull[2].findAll("td")
                chk["Name"] = name[1].text[1:]
                if(len(name[1].text[1:]) > NameLength):
                    NameLength = len(name[1].text[1:])
                sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
                sgpa_1 = sgpa.findAll("tr")
                for l in range(2,len(sgpa_1)):
                    check = sgpa_1[l].findAll("td")
                    if(Sem_var in check[0].text):
                        txt = check[1].text
                        if "   PROMOTED-- " in txt:
                            txt="-"
                        elif "   PROMOTED-" in txt:
                            txt = txt.replace("   PROMOTED-", "")
                        elif "   DETAINED " in txt:
                            txt = txt.replace(" ","")
                        elif "  PASSED-" in txt:
                            txt = txt.replace("  PASSED-","")
                            txt = txt.replace(" ","")
                        chk["SGPA"] = txt
                
                conts = container.findAll("tr")
                chks=0
                for j in conts:
                    chks+=1
                    if chks>2:
                        self.Subs = j.findAll("td")
                        sub = self.Subs[1].text.replace("/xa0","")
                        if(sub not in self.Subss):
                            self.Subss.append(sub)
                        chk[sub] = self.Subs[3].text.replace("/xa0","")
            else:
                driver.get(Url)
                element = driver.find_elements(By.TAG_NAME, "input")
                if len(element) == 3:
                    element[1].send_keys(q)
                    element[2].click()
                else:
                    element[2].send_keys(q)
                    element[3].click()
                html = driver.page_source
                pageSoup = soup(html,"html.parser")
                container = pageSoup.find("table",{"id":"AutoNumber4"})

                if container != None:
                    containerName = pageSoup.find("table",{"id":"AutoNumber3"})
                    nameFull = containerName.findAll("tr")
                    name = nameFull[2].findAll("td")
                    chk["Name"] = name[1].text[1:]
                    if(len(name[1].text[1:]) > NameLength):
                        NameLength == len(name[1].text[1:])
                    
                    sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
                    sgpa_1 = sgpa.findAll("tr")
                    for l in range(2,len(sgpa_1)):
                        check = sgpa_1[l].findAll("td")
                        if(Sem_var in check[0].text):
                            txt = check[1].text
                            if "   PROMOTED-- " in txt:
                                txt="-"
                            elif "   PROMOTED-" in txt:
                                txt = txt.replace("   PROMOTED-", "")
                            elif "   DETAINED " in txt:
                                txt = txt.replace(" ","")
                            elif "  PASSED-" in txt:
                                txt = txt.replace("  PASSED-","")
                                txt = txt.replace(" ","")
                            chk["SGPA"] = txt
                    
                    conts = container.findAll("tr")
                    chks=0
                    for j in conts:
                        chks+=1
                        if chks>2:
                            self.Subs = j.findAll("td")
                            sub = self.Subs[1].text.replace("/xa0","")
                            if(sub not in self.Subss):
                                self.Subss.append(sub)
                            chk[sub] = self.Subs[3].text.replace("/xa0","")
                driver.get(UrlRv)
            if chk != None:
                self.outs[x] = chk
                chk = {}
        self.NameLengths.append(NameLength)
        driver.quit()

    def multiThread_RV(self, x, driver, NameLength, chk): 
##        NameLength = 0
##        chk = {}
        try:
            if(soup(driver.page_source,"html.parser").findAll("h1")[0].text == "HTTP Status 500 – Internal Server Error"):
                return False
            if soup(driver.page_source,"html.parser").find("div", {"id":"main-message"}).findAll("span")[0].text == "Your connection was interrupted":
                return False
        except:
            pass
        try:
            element = driver.find_elements(By.TAG_NAME, "input")
            if len(element) == 3:
                element[1].send_keys(x)
                element[2].click()
            else:
                element[2].send_keys(x)
                element[3].click()
            
            html = driver.page_source
            pageSoup = soup(html,"html.parser")
            container = pageSoup.find("table",{"id":"AutoNumber4"})
            check = pageSoup.find("table",{"id":"AutoNumber1"})
            if check == None:
                return False
            
            if check.findAll("b")[1].text == "Personal Details":
                containerName = pageSoup.find("table",{"id":"AutoNumber3"})
                nameFull = containerName.findAll("tr")
                name = nameFull[2].findAll("td")
                chk["Name"] = name[1].text[1:]
                if(len(name[1].text[1:]) > NameLength):
                    NameLength = len(name[1].text[1:])
                sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
                sgpa_1 = sgpa.findAll("tr")
                for l in range(2,len(sgpa_1)):
                    check = sgpa_1[l].findAll("td")
                    if(Sem_var in check[0].text):
                        txt = check[1].text
                        if "   PROMOTED-- " in txt:
                            txt="-"
                        elif "   PROMOTED-" in txt:
                            txt = txt.replace("   PROMOTED-", "")
                        elif "   DETAINED " in txt:
                            txt = txt.replace(" ","")
                        elif "  PASSED-" in txt:
                            txt = txt.replace("  PASSED-","").replace(" ","")
                        chk["SGPA"] = txt        
                conts = container.findAll("tr")
                chks=0
                for j in conts:
                    chks+=1
                    if chks>2:
                        Subs = j.findAll("td")
                        sub = Subs[1].text.replace("/xa0","")
                        if(sub not in self.Subss):
                            self.Subss.append(sub)
                        chk[sub] = Subs[3].text.replace("/xa0","")
            elif "The Hall Ticket Number" == pageSoup.find("table",{"id":"AutoNumber1"}).findAll("b")[1].text[9:31]:
                return self.multiThread(x, driver, NameLength, chk)
            else:
                return False

            if chk != None:
                self.outs[x] = chk
                chk = {}
            self.NameLengths.append(NameLength)
            return True
        except Exception as e:
            return False

    def multiThread(self, x, driver, NameLength, chk, val): 
##        NameLength = 0
##        chk = {}
        try:
            if(soup(driver.page_source,"html.parser").findAll("h1")[0].text == "HTTP Status 500 – Internal Server Error"):
                return False
            if soup(driver.page_source,"html.parser").find("div", {"id":"main-message"}).findAll("span")[0].text == "Your connection was interrupted":
                return False
        except:
            pass
        try:
            if(val == 0):
                element = driver.find_elements(By.TAG_NAME, "input")
                if len(element) == 3:
                    element[1].send_keys(x)
                    element[2].click()
                else:
                    element[2].send_keys(x)
                    element[3].click()
            
            html = driver.page_source
            pageSoup = soup(html,"html.parser")
            container = pageSoup.find("table",{"id":"AutoNumber4"})
            check = pageSoup.find("table",{"id":"AutoNumber1"})
            if check == None:
                return False
            
            if check.findAll("b")[1].text == "Personal Details":
                containerName = pageSoup.find("table",{"id":"AutoNumber3"})
                nameFull = containerName.findAll("tr")
                name = nameFull[2].findAll("td")
                chk["Name"] = name[1].text[1:]
                if(len(name[1].text[1:]) > NameLength):
                    NameLength = len(name[1].text[1:])
                sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
                sgpa_1 = sgpa.findAll("tr")
                for l in range(2,len(sgpa_1)):
                    check = sgpa_1[l].findAll("td")
                    if(Sem_var in check[0].text):
                        txt = check[1].text
                        if "   PROMOTED-- " in txt:
                            txt="-"
                        elif "   PROMOTED-" in txt:
                            txt = txt.replace("   PROMOTED-", "")
                        elif "   DETAINED " in txt:
                            txt = txt.replace(" ","")
                        elif "  PASSED-" in txt:
                            txt = txt.replace("  PASSED-","").replace(" ","")
                        chk["SGPA"] = txt        
                conts = container.findAll("tr")
                chks=0
                for j in conts:
                    chks+=1
                    if chks>2:
                        Subs = j.findAll("td")
                        sub = Subs[1].text.replace("/xa0","")
                        if(sub not in self.Subss):
                            self.Subss.append(sub)
                        chk[sub] = Subs[3].text.replace("/xa0","")
            elif "The Hall Ticket Number" == pageSoup.find("table",{"id":"AutoNumber1"}).findAll("b")[1].text[9:31]:
                return True
            else:
                return False

            if chk != None:
                self.outs[x] = chk
                chk = {}
            self.NameLengths.append(NameLength)
            return True
        except Exception as e:
            return False
        
    def multiThreadCall(self, arr):
        s=Service(DriverLocation)
        driver = webdriver.Chrome(service=s)
        driver.get(Url)
        for x in arr:
            while(self.multiThread(x, driver, 0, {}, 0) == False):
                driver.refresh()
                self.multiThread(x, driver, 0, {}, 1)
        driver.quit()
    def multiThreadRVCall(self, arr):
        s=Service(DriverLocation)
        driver = webdriver.Chrome(service=s)
        driver.get(Url)
        for x in arr:
            while(self.multiThread_RV(x, driver, 0, {}) == False):
                driver.refresh()
                self.multiThread_RV(x, driver, 0, {})
        driver.quit()

    def startScrapping(self):
        self.startReading()
        for i in range(len(self.rollNos)):
            chk = int(len(self.rollNos[i])*0.25)
            start = 0
            end = chk
            Threads = []
            if(UrlRv != None):
                for x in range(3):
                    Threads.append(threading.Thread(target = self.multiThreadRVCall, args=(self.rollNos[i][start:end+1],)))
                    start = end+1
                    end += chk
                Threads.append(threading.Thread(target = self.multiThreadRVCall, args=(self.rollNos[i][start:len(self.rollNos[i])],)))    
                for j in Threads:
                    j.start()
                for j in Threads:
                    j.join()
                self.Subs.append(self.Subss)
                self.Subss = []
            else:
                for x in range(3):
                    Threads.append(threading.Thread(target = self.multiThreadCall, args=(self.rollNos[i][start:end+1],)))
                    start = end+1
                    end += chk
                Threads.append(threading.Thread(target = self.multiThreadCall, args=(self.rollNos[i][start:len(self.rollNos[i])],)))    
                for j in Threads:
                    j.start()
                for j in Threads:
                    j.join()
                self.Subs.append(self.Subss)
                self.Subss = []
        self.startSaving()

    def startSaving(self):
        global alignment
        alignment = Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
        wb1 = openpyxl.load_workbook(Input)
        sheets = wb1.sheetnames
        for i in range(len(sheets)):
            wb1.active = i
            sheet1 = wb1.active
            sheet1['A1'].value="Roll No."
            sheet1['B1'].value="Student Name"
            sheet1['C1'].value="SGPA"
            sheet1.column_dimensions['A'].width=14
            sheet1.column_dimensions['B'].width=self.NameLengths[i]+15
            sheet1.column_dimensions['C'].width=10
            sheet1['A1'].font = Font(bold=True)
            sheet1['B1'].font = Font(bold=True)
            sheet1['C1'].font = Font(bold=True)
            sheet1['A1'].alignment=alignment
            sheet1['B1'].alignment=alignment
            sheet1['C1'].alignment=alignment
            letter1 = 68
            for j in self.Subs[i]:
                if(letter1 <= 90):
                    temp = chr(letter1)+str(1)
                    sheet1[temp].value=j
                    sheet1.column_dimensions[temp.replace("1", "")].width = len(j)+4
                    sheet1[temp].alignment = alignment
                    sheet1[temp].font = Font(bold=True)
                    letter1+=1
                elif(letter1 > 90):
                    temp = 'A' + chr(letter1-26) + str(1)
                    sheet1[temp].value=j
                    sheet1.column_dimensions[temp.replace("1", "")].width = len(j)+4
                    sheet1[temp].alignment = alignment
                    sheet1[temp].font = Font(bold=True)
                    letter1+=1
                elif(letter1 > 116):
                    temp = 'B' + chr(letter1-52) + str(1)
                    sheet1[temp].value=j
                    sheet1.column_dimensions[temp.replace("1", "")].width = len(j)+4
                    sheet1[temp].alignment = alignment
                    sheet1[temp].font = Font(bold=True)
                    letter1+=1
                elif(letter1 > 142):
                    temp = 'B' + chr(letter1-78) + str(1)
                    sheet1[temp].value=j
                    sheet1.column_dimensions[temp.replace("1", "")].width = len(j)+4
                    sheet1[temp].alignment = alignment
                    sheet1[temp].font = Font(bold=True)
                    letter1+=1
                
            row = sheet1.max_row
            for q in range(2,row+1):
                letter2 = 68
                temp = self.outs.get(sheet1['A'+str(q)].value)
                if(temp == None):
                    continue
                sheet1["B"+str(q)].value = temp.get("Name")
                sheet1["C"+str(q)].value = temp.get("SGPA")
                sheet1["B"+str(q)].alignment = alignment
                sheet1["C"+str(q)].alignment = alignment
                sheet1["B"+str(q)].font = Font(bold=True)
                sheet1["C"+str(q)].font = Font(bold=True)
                for w in range(len(self.Subs[i])):
                    if(letter2 <= 90):
                        temp1 = chr(letter2)+str(q)
                        if temp.get(self.Subs[i][w]) != None:
                            sheet1[temp1] = temp.get(self.Subs[i][w])
                            sheet1[temp1].alignment = alignment
                        else:
                            sheet1[temp1] = "-"
                            sheet1[temp1].alignment = alignment
                        letter2+=1
                    elif(letter2 > 90):
                        temp1 = 'A' + chr(letter2-26) + str(q)
                        if temp.get(self.Subs[i][w]) != None:
                            sheet1[temp1] = temp.get(self.Subs[i][w])
                            sheet1[temp1].alignment = alignment
                        else:
                            sheet1[temp1] = "-"
                            sheet1[temp1].alignment = alignment
                        letter2+=1
                    elif(letter2 > 116):
                        temp1 = 'B' + chr(letter2-52) + str(q)
                        if temp.get(self.Subs[i][w]) != None:
                            sheet1[temp1] = temp.get(self.Subs[i][w])
                            sheet1[temp1].alignment = alignment
                        else:
                            sheet1[temp1] = "-"
                            sheet1[temp1].alignment = alignment
                        letter2+=1
                    elif(letter2 > 142):
                        temp1 = 'C' + chr(letter2-78) + str(q)
                        if temp.get(self.Subs[i][w]) != None:
                            sheet1[temp1] = temp.get(self.Subs[i][w])
                            sheet1[temp1].alignment = alignment
                        else:
                            sheet1[temp1] = "-"
                            sheet1[temp1].alignment = alignment
                        letter2+=1
        wb1.active=0
        wb1.save(Output)

    def tkint(self):    
        window = tkinter.Tk()
        window.title("OU Results")
        window.geometry('593x320')

        tkinter.Label(window,text="WebDriver Location : ").place(x=10,y=43)
        tkinter.Label(window,text="Input File Name : ").place(x=32,y=73)
        tkinter.Label(window,text="Output File Name : ").place(x=25,y=103)
        txt_3 = tkinter.Entry(window,width=40)
        txt_3.place(x=160,y=45,height=26)
        txt_1 = tkinter.Entry(window,width=40)
        txt_1.place(x=160,y=75,height=26)
        txt_2 = tkinter.Entry(window,width=40)
        txt_2.place(x=160,y=105,height=26)

        tkinter.Label(window,text="Results URL : ").place(x=32,y=160)
        tkinter.Label(window,text="Results RV URL : ").place(x=25,y=186)
        URL = tkinter.Entry(window,width=40)
        URL.place(x=160,y=160,height=26)
        URL_RV = tkinter.Entry(window,width=40)
        URL_RV.place(x=160,y=190,height=26)

        var1 = tkinter.IntVar()

        def clicked():
            global Input
            Input = txt_1.get()
            #Input = "D:/Prajwal/Python/Web_scraping/4th yr.xlsx"
            global Output
            Output = txt_2.get()
            #Output = "D:/Prajwal/Python/Out.xlsx"
            global Url
            Url = URL.get()
            #Url = "https://www.osmania.ac.in/res07/20221206.jsp"
            #global DriverLocation
            #DriverLocation = "C:/Users/prajw/Downloads/chromedriver_win32/chromedriver.exe"
            global UrlRv
            UrlRv = None
            if len(URL_RV.get()) != 0:
                UrlRv = URL_RV.get()
            window.destroy()

        def get_sem(selection):
            global Sem_var
            Sem_var = selection

        def browse_In():
            file=tkinter.filedialog.askopenfilename()
            global Input
            txt_1.delete(0, END)
            txt_1.insert(tkinter.END, file)
            Input = txt_1.get()

        def browse_Out():
            file = tkinter.filedialog.askdirectory()
            global Output
            txt_2.delete(0, END)
            txt_2.insert(tkinter.END, file+"/Output.xlsx")
            Output = txt_2.get()

        def driver():
            file=tkinter.filedialog.askopenfilename()
            global DriverLocation
            txt_3.delete(0, END)
            txt_3.insert(tkinter.END, file)
            DriverLocation = txt_3.get()
            #DriverLocation = "C:/Users/prajw/Downloads/chromedriver_win32/chromedriver.exe"



        variable = StringVar(window)
        tkinter.Label(window,text="Semester : ").place(x=160,y=250)
        sem_var=OptionMenu(window,variable,"1","2","3","4","5","6","7","8",command=get_sem)
        sem_var.pack()
        sem_var.place(x=240,y=244)

        bt_3=tkinter.Button(window,text="Browse",command=driver)
        bt_3.place(x=500, y=45,height=26)    
        bt_1=tkinter.Button(window,text="Browse",command=browse_In)
        bt_1.place(x=500, y=75,height=26)
        bt_2=tkinter.Button(window,text="Browse",command=browse_Out)
        bt_2.place(x=500, y=105,height=26)
        #ch_1=tkinter.Checkbutton(window,text="Marks",variable=var1).place(x=218,y=130)
        #print(var1)
        bt=tkinter.Button(window,text="Submit",command=clicked,width=8).place(x=320,y=245)
        window.mainloop()
        #Tkinter ______End_______
sc = Scrape()
sc.tkint()
sc.startScrapping()
