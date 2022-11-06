import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import threading
import time

from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as uReq

from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver import Chrome

import os

wb1 = openpyxl.load_workbook("D:/Prajwal/Python/Web_scraping/Input.xlsx")
sheets = wb1.sheetnames

rollNos = []   

for i in range(len(sheets)):
    wb1.active = i
    sheet1 = wb1.active
    row = sheet1.max_row
    roll=[]
    for i in range(2,row+1):
        roll.append(sheet1.cell(i,1).value)
    rollNos.append(roll)

outs = {}
NameLengths = []
subss = []

def multiThread(i, start, end,):
    NameLength = 0
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
    driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)
    s=Service("C:/Users/prajw/Downloads/chromedriver_win32/chromedriver.exe")
    driver = webdriver.Chrome(service=s)
    driver.get('https://www.osmania.ac.in/res07/20220875.jsp')
    chk = {}
    for q in range(start,end):        
        element = driver.find_elements(By.TAG_NAME, "input")
        if len(element) == 3:
            element[1].send_keys(rollNos[i][q])
            element[2].click()
        else:
            element[2].send_keys(rollNos[i][q])
            element[3].click()
        
        html = driver.page_source
        pageSoup = soup(html,"html.parser")
        container = pageSoup.find("table",{"id":"AutoNumber4"})
        
        if container != None:
            containerName = pageSoup.find("table",{"id":"AutoNumber3"})
            nameFull = containerName.findAll("tr")
            name = nameFull[2].findAll("td")
            chk["Name"] = name[1].text[1:]
            if(len(name[1].text[1:]) > NameLength):
                NameLength = len(name[1].text[1:])
            sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
            sgpa_1 = sgpa.findAll("tr")
            for l in range(2,len(sgpa_1)):
                check = sgpa_1[l].findAll("td")
                if("3" in check[0].text):
                    txt = check[1].text
                    if "   PROMOTED-- " in txt:
                        txt="-"
                    elif "  PASSED-" in txt:
                        txt = txt.replace("  PASSED-","")
                        txt = txt.replace(" ","")
                    chk["SGPA"] = txt
            
            conts = container.findAll("tr")
            chks=0
            for j in conts:
                chks+=1
                if chks>2:
                    subs = j.findAll("td")
                    sub = subs[1].text.replace("\xa0","")
                    if(sub not in subss):
                        subss.append(sub)
                    chk[sub] = subs[3].text.replace("\xa0","")
        else:
            driver.get('https://www.osmania.ac.in/res07/20220625.jsp')
            element = driver.find_elements(By.TAG_NAME, "input")
            if len(element) == 3:
                element[1].send_keys(rollNos[i][q])
                element[2].click()
            else:
                element[2].send_keys(rollNos[i][q])
                element[3].click()
            html = driver.page_source
            pageSoup = soup(html,"html.parser")
            container = pageSoup.find("table",{"id":"AutoNumber4"})

            if container != None:
                containerName = pageSoup.find("table",{"id":"AutoNumber3"})
                nameFull = containerName.findAll("tr")
                name = nameFull[2].findAll("td")
                chk["Name"] = name[1].text[1:]
                if(len(name[1].text[1:]) > NameLength):
                    NameLength == len(name[1].text[1:])
                
                sgpa = pageSoup.find("table",{"id":"AutoNumber5"})
                sgpa_1 = sgpa.findAll("tr")
                for l in range(2,len(sgpa_1)):
                    check = sgpa_1[l].findAll("td")
                    if("3" in check[0].text):
                        txt = check[1].text
                        if "   PROMOTED-- " in txt:
                            txt="-"
                        elif "  PASSED-" in txt:
                            txt = txt.replace("  PASSED-","")
                            txt = txt.replace(" ","")
                        chk["SGPA"] = txt
                
                conts = container.findAll("tr")
                chks=0
                for j in conts:
                    chks+=1
                    if chks>2:
                        subs = j.findAll("td")
                        sub = subs[1].text.replace("\xa0","")
                        if(sub not in subss):
                            subss.append(sub)
                        chk[sub] = subs[3].text.replace("\xa0","")
            driver.get('https://www.osmania.ac.in/res07/20220875.jsp')
        if chk != None:
            outs[rollNos[i][q]] = chk
            chk = {}
    NameLengths.append(NameLength)
    driver.quit()

Subs = []
for i in range(len(rollNos)):
    chk = int(len(rollNos[i])*0.25)
    start = 0
    end = chk
    Threads = []
    for x in range(3):
        Threads.append(threading.Thread(target = multiThread, args=(i, start, end+1,)))
        start = end+1
        end += chk
    Threads.append(threading.Thread(target = multiThread, args=(i, start, len(rollNos[i]),)))    
    for j in Threads:
        j.start()
    for j in Threads:
        j.join()
    Subs.append(subss)
    subss = []

#________________________________________#
#________________________________________#
alignment = Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)

for i in range(len(sheets)):
    wb1.active = i
    sheet1 = wb1.active
    sheet1['A1'].value="Roll No."
    sheet1['B1'].value="Student Name"
    sheet1['C1'].value="SGPA"
    sheet1.column_dimensions['A'].width=14
    sheet1.column_dimensions['B'].width=NameLengths[i]+15
    sheet1.column_dimensions['C'].width=10
    sheet1['A1'].font = Font(bold=True)
    sheet1['B1'].font = Font(bold=True)
    sheet1['C1'].font = Font(bold=True)
    sheet1['A1'].alignment=alignment
    sheet1['B1'].alignment=alignment
    sheet1['C1'].alignment=alignment
    letter1 = 68
    for j in Subs[i]:
        sheet1[chr(letter1)+str(1)].value=j
        sheet1.column_dimensions[chr(letter1)].width = len(j)+4
        sheet1[chr(letter1)+"1"].alignment = alignment
        sheet1[chr(letter1)+"1"].font = Font(bold=True)
        letter1+=1
    row = sheet1.max_row
    for q in range(2,row+1):
        letter2 = 68
        temp = outs.get(sheet1['A'+str(q)].value)
        sheet1["B"+str(q)].value = temp.get("Name")
        sheet1["C"+str(q)].value = temp.get("SGPA")
        sheet1["B"+str(q)].alignment = alignment
        sheet1["C"+str(q)].alignment = alignment
        sheet1["B"+str(q)].font = Font(bold=True)
        sheet1["C"+str(q)].font = Font(bold=True)
        for w in range(len(Subs[i])):
            if temp.get(Subs[i][w]) != None:
                sheet1[chr(letter2)+str(q)] = temp.get(Subs[i][w])
                sheet1[chr(letter2)+str(q)].alignment = alignment
            else:
                sheet1[chr(letter2)+str(q)] = "-"
                sheet1[chr(letter2)+str(q)].alignment = alignment
            letter2+=1
wb1.active=0
wb1.save("D:\Prajwal\Python\Web_scraping\Output_Retry.xlsx")
