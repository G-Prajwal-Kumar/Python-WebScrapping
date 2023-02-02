from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as uReq
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
alignment = Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
alignment_1 = Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)


my_url = 'https://www.task.telangana.gov.in/Registered-Colleges'

uClient = uReq(my_url)
page_html = uClient.read()
page_soup = soup(page_html,"html.parser")

containers = page_soup.findAll("div",{"class":"table-responsive"})


container=containers[1]
clg_type=[]
dists_no=[]
dist_value=[]
principal=[]
placement_officer=[]
k=1
count=0

c=0
d=0
e=0


exp1=container.table.tbody
exp2=exp1.findAll("tr")

for block in container.table.thead.tr:
    if block != '\n':
        clg_type.append(block.text)

for block in exp2:
    exp3=block.find("th")
    print(" ")
    print(exp3.text)
    exp4=block.findAll("td")
    count=0
    count_2=0
    
    wb=Workbook()
    if (exp3.text) not in ["Total","Medchal","Ranga Reddy","Sangareddy","Hyderabad","Warangal urban","Nizamabad","Khammam","Karimnagar"] :
        for u in exp4:
            count+=1
            count_3=2
            count_2+=1
            print(count_2)
            if(count<7):
                if int(u.text) != 0 and str(u.a["class"]) != "reg_clg_num":
                    my_url_new=u.a["href"]
                    uClient_new = uReq(my_url_new)
                    page_html_new = uClient_new.read()
                    page_soup_new = soup(page_html_new,"html.parser")
                    containers_new = page_soup_new.findAll("table",{"class":"table"})
                    exp5=containers_new[1].tbody.findAll("tr")
                    sector = page_soup_new.find("h1",{"class":"entry-title pbot10 animated fadeInLeft"})
                    sheet_name=sector.text
                    sheet_name=sheet_name.replace(" Colleges","")
                    print(sheet_name)
                    '''
                    for sheet in wb.worksheets:
                        print(sheet)
                        if sheet == wb.worksheets[0]:
                            wb.create_sheet(title=
                        else:
                            wb.create_sheet(title=sheet_name)
                    '''
                    print(count_2)
                    if count_2==1:
                        ss_sheet = wb['Sheet']
                        ss_sheet.title = (exp3.text)+" "+sheet_name
                    else:
                        wb.create_sheet(title=sheet_name)
                    wb.active=count_2-1
                    sh1=wb.active
                    sh1['A1'].value="SI"
                    sh1['B1'].value="COLLEGE NAME"
                    sh1['C1'].value="STREAM"
                    sh1['D1'].value="DESIGNATION"
                    sh1['E1'].value="CONTACT DETAILS"
                    sh1['F1'].value="ADDRESS"
                    sh1.column_dimensions['A'].width = 5
                    

                    for i in ['A','B','C','D','E','F']:
                        
                        currentCell=sh1[i+'1']
                        currentCell.alignment = alignment_1
                        sh1[i+'1'].fill=PatternFill("solid",fgColor="9bbb59")
                        
                        sh1[i+'1'].font = Font(bold=True)
                    
                
                    for h in exp5:
                        a=h.findAll("a")
                        for w in a:
                            if (w.text) != "":
                                    
                                my_url_new_1=w['href']
                                uClient_new_1 = uReq(my_url_new_1)
                                page_html_new_1 = uClient_new_1.read()
                                page_soup_new_1 = soup(page_html_new_1,"html.parser")
                                containers_new_1 = page_soup_new_1.findAll("td",{"class":"student-details"})
                                college_name = containers_new_1[0].text
                                college_address = containers_new_1[5].text
                                print(college_name,college_address)


                                sh1['A'+str(count_3)].value=count_3-1
                                sh1['B'+str(count_3)].value=college_name
                                sh1['C'+str(count_3)].value=sheet_name
                                sh1['F'+str(count_3)].value=college_address

                                if len(college_name)>d:
                                    d=len(college_name)
                                if len(college_address)>e:
                                    e=len(college_address)
                                sh1.column_dimensions['B'].width = d+4
                                sh1.column_dimensions['C'].width = len(sheet_name)+4
                                sh1.column_dimensions['F'].width = e+4
                                

                                containers_new_1_2 = page_soup_new_1.findAll("div",{"class":"table-responsive"})
                                exp6=containers_new_1_2[1].tbody.findAll("tr")
                                
                                exp6_a=exp6[0].findAll("td")
                                exp6_b=exp6[1].findAll("td")
                                for p in exp6_a:
                                    principal.append(p.text)
                                for o in exp6_b:
                                    placement_officer.append(o.text)

                                print(principal)
                                print(placement_officer)
                                '''
                                cell_format = wb.add_format({'text_wrap': True})
                                sh1.write('D'+str(count_3),principal[0]+"\n"+placement_officer[0],cell_format)
                                sh1.write('E'+str(count_3),principal[1]+"  "+principal[2]+"\n"+placement_officer[1]+"  "+placement_officer[2],cell_format)
                                
                                '''
                                for i in ['A','B','C','D','E','F']:
                                    sh1[i+str(count_3)].alignment = alignment
                                a=len(principal[1]+' , '+principal[2])
                                b=len(placement_officer[1]+' , '+placement_officer[2])
                                if a>b:
                                    c=a
                                else:
                                    c=b
                                sh1.column_dimensions['E'].width = c
                                sh1.column_dimensions['D'].width = 20
                                sh1['D'+str(count_3)].value=principal[0]+"\n"+placement_officer[0]
                                sh1['E'+str(count_3)].value=principal[1]+' , '+principal[2]+"\n"+placement_officer[1]+' , '+placement_officer[2]

                                
                                principal=[]
                                placement_officer=[]
                                print(count_3)
                                count_3+=1
                else:
                    count_2-=1
                    
    wb.active=0
    wb.save(exp3.text+".xlsx")
    
